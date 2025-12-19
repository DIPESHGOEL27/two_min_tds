"""
Excel report generation for TDS Challan data.
Produces Excel files matching the required schema with data and summary sheets.
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from models import ChallanRecord, ValidationStatus

logger = logging.getLogger(__name__)


# Excel column schema - defines the exact column order and headers
EXCEL_COLUMNS = [
    "TAN",
    "Deductor Name",
    "Assessment Year",
    "Financial Year",
    "Major Head",
    "Minor Head",
    "Nature of Payment",
    "Total Amount",
    "Amount in Words",
    "CIN",
    "BSR Code",
    "Challan No",
    "Date of Deposit",
    "Bank Name",
    "Bank Ref No",
    "Tax_A",
    "Tax_B",
    "Tax_C",
    "Tax_D",
    "Tax_E",
    "Tax_F",
    "Source File",
    "Row Confidence",
    "Validation Flag",
    "Notes",
]

# Column data types for documentation
COLUMN_TYPES = {
    "TAN": "string",
    "Deductor Name": "string",
    "Assessment Year": "string",
    "Financial Year": "string",
    "Major Head": "string",
    "Minor Head": "string",
    "Nature of Payment": "string",
    "Total Amount": "float",
    "Amount in Words": "string",
    "CIN": "string",
    "BSR Code": "string",
    "Challan No": "string",
    "Date of Deposit": "date (ISO format YYYY-MM-DD)",
    "Bank Name": "string",
    "Bank Ref No": "string",
    "Tax_A": "float",
    "Tax_B": "float",
    "Tax_C": "float",
    "Tax_D": "float",
    "Tax_E": "float",
    "Tax_F": "float",
    "Source File": "string",
    "Row Confidence": "float (0-1)",
    "Validation Flag": "string (OK/FLAG)",
    "Notes": "string",
}


class ExcelWriter:
    """Generate Excel reports from extracted challan records."""

    def __init__(self):
        self.columns = EXCEL_COLUMNS

        # Styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.flag_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def write(
        self,
        records: List[ChallanRecord],
        output_path: Path,
        include_summary: bool = True
    ) -> Path:
        """
        Write records to Excel file.

        Args:
            records: List of ChallanRecord objects
            output_path: Path for output Excel file
            include_summary: Whether to include summary sheet

        Returns:
            Path to created Excel file
        """
        logger.info(f"Writing {len(records)} records to {output_path}")

        # Create workbook
        wb = Workbook()

        # Main data sheet
        ws_data = wb.active
        ws_data.title = "TDS Challans"
        self._write_data_sheet(ws_data, records)

        # Summary sheet
        if include_summary and records:
            ws_summary = wb.create_sheet("Summary")
            self._write_summary_sheet(ws_summary, records)

        # Flagged records sheet
        flagged = [r for r in records if r.validation_flag == ValidationStatus.FLAG]
        if flagged:
            ws_flagged = wb.create_sheet("Flagged Records")
            self._write_data_sheet(ws_flagged, flagged)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Excel file saved: {output_path}")
        return output_path

    def _write_data_sheet(self, ws, records: List[ChallanRecord]):
        """Write main data to worksheet."""
        # Write headers
        for col_idx, header in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        # Write data rows
        for row_idx, record in enumerate(records, 2):
            row_data = record.to_excel_row()

            for col_idx, header in enumerate(self.columns, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border

                # Format numbers
                if header in ("Total Amount", "Tax_A", "Tax_B", "Tax_C", "Tax_D", "Tax_E", "Tax_F"):
                    cell.number_format = "#,##0.00"
                elif header == "Row Confidence":
                    cell.number_format = "0.00%"

                # Highlight flagged rows
                if header == "Validation Flag":
                    if value == "FLAG":
                        cell.fill = self.flag_fill
                    else:
                        cell.fill = self.ok_fill

        # Adjust column widths
        self._auto_fit_columns(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _write_summary_sheet(self, ws, records: List[ChallanRecord]):
        """Write summary statistics to worksheet."""
        # Title
        ws["A1"] = "TDS Challan Processing Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Generation timestamp
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Overall statistics
        ws["A4"] = "Overall Statistics"
        ws["A4"].font = Font(bold=True)

        stats = [
            ("Total Records", len(records)),
            ("OK Records", sum(1 for r in records if r.validation_flag == ValidationStatus.OK)),
            ("Flagged Records", sum(1 for r in records if r.validation_flag == ValidationStatus.FLAG)),
            ("Total Amount (Sum)", sum(r.total_amount or 0 for r in records)),
            ("Average Confidence", sum(r.row_confidence for r in records) / len(records) if records else 0),
        ]

        for idx, (label, value) in enumerate(stats, 5):
            ws[f"A{idx}"] = label
            ws[f"B{idx}"] = value
            if label == "Total Amount (Sum)":
                ws[f"B{idx}"].number_format = "#,##0.00"
            elif label == "Average Confidence":
                ws[f"B{idx}"].number_format = "0.00%"

        # Summary by TAN
        ws["A11"] = "Summary by TAN"
        ws["A11"].font = Font(bold=True)

        ws["A12"] = "TAN"
        ws["B12"] = "Record Count"
        ws["C12"] = "Total Amount"
        ws["D12"] = "Flagged"

        for cell in ["A12", "B12", "C12", "D12"]:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = self.header_fill
            ws[cell].font = self.header_font

        # Group by TAN
        tan_summary = {}
        for record in records:
            tan = record.tan or "Unknown"
            if tan not in tan_summary:
                tan_summary[tan] = {"count": 0, "amount": 0, "flagged": 0}
            tan_summary[tan]["count"] += 1
            tan_summary[tan]["amount"] += record.total_amount or 0
            if record.validation_flag == ValidationStatus.FLAG:
                tan_summary[tan]["flagged"] += 1

        row = 13
        for tan, data in sorted(tan_summary.items()):
            ws[f"A{row}"] = tan
            ws[f"B{row}"] = data["count"]
            ws[f"C{row}"] = data["amount"]
            ws[f"C{row}"].number_format = "#,##0.00"
            ws[f"D{row}"] = data["flagged"]
            row += 1

        # List of flagged records
        flagged = [r for r in records if r.validation_flag == ValidationStatus.FLAG]
        if flagged:
            row += 2
            ws[f"A{row}"] = "Flagged Records Details"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            ws[f"A{row}"] = "Source File"
            ws[f"B{row}"] = "CIN"
            ws[f"C{row}"] = "Amount"
            ws[f"D{row}"] = "Issue"

            for cell in [f"A{row}", f"B{row}", f"C{row}", f"D{row}"]:
                ws[cell].font = Font(bold=True)

            row += 1
            for record in flagged:
                ws[f"A{row}"] = record.source_file
                ws[f"B{row}"] = record.cin
                ws[f"C{row}"] = record.total_amount
                ws[f"C{row}"].number_format = "#,##0.00"
                ws[f"D{row}"] = record.notes
                row += 1

        self._auto_fit_columns(ws)

    def _auto_fit_columns(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-fit column widths based on content."""
        from openpyxl.cell.cell import MergedCell
        
        for column_cells in ws.columns:
            max_length = 0
            column = None
            
            # Find first non-merged cell to get column letter
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column = cell.column_letter
                    break
            
            if column is None:
                continue  # Skip if entire column is merged

            for cell in column_cells:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width


def write_excel(
    records: List[ChallanRecord],
    output_path: Path,
    include_summary: bool = True
) -> Path:
    """Convenience function to write Excel file."""
    writer = ExcelWriter()
    return writer.write(records, output_path, include_summary)


def get_column_schema() -> Dict[str, str]:
    """Get the Excel column schema with types."""
    return COLUMN_TYPES.copy()


def get_sample_row() -> Dict[str, Any]:
    """Get a sample row for documentation."""
    return {
        "TAN": "BLRS05586H",
        "Deductor Name": "SYAMBHAVAN FOODS LLP",
        "Assessment Year": "2026-27",
        "Financial Year": "2025-26",
        "Major Head": "Corporation Tax (0020)",
        "Minor Head": "TDS/TCS Payable by Taxpayer (200)",
        "Nature of Payment": "94J",
        "Total Amount": 19395.00,
        "Amount in Words": "Rupees Nineteen Thousand Three Hundred And Ninety Five Only",
        "CIN": "25100700517216HDFC",
        "BSR Code": "0510016",
        "Challan No": "12866",
        "Date of Deposit": "2025-10-07",
        "Bank Name": "HDFC Bank",
        "Bank Ref No": "N2528040495795",
        "Tax_A": 19395.00,
        "Tax_B": 0.00,
        "Tax_C": 0.00,
        "Tax_D": 0.00,
        "Tax_E": 0.00,
        "Tax_F": 0.00,
        "Source File": "25100700517216HDFC_ChallanReceipt- Input Command Challan.pdf",
        "Row Confidence": 0.95,
        "Validation Flag": "OK",
        "Notes": "",
    }
